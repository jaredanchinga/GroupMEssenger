import pandas as pd
from openpyxl import load_workbook
import logging
import os

class ExcelHandler:
    def __init__(self, excel_path='groupme.xlsx'):
        self.excel_path = excel_path
        # Check if file exists on initialization
        if not os.path.exists(self.excel_path):
            logging.error(f"Excel file not found: {self.excel_path}")
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
        logging.info(f"Excel file found: {self.excel_path}")

    def cut_link_to_sheet(self, link, to_sheet):
        """Cut (not copy) a link from links sheet to destination sheet"""
        wb = load_workbook(self.excel_path)
        
        try:
            # Get source and destination sheets
            source_sheet = wb['links']  # Always cut from links sheet
            dest_sheet = wb[to_sheet]
            
            # Find link in source sheet
            link_row = None
            for row in range(1, source_sheet.max_row + 1):
                cell_value = source_sheet.cell(row=row, column=1).value
                if cell_value == link:
                    link_row = row
                    break
                
            if link_row:
                # Add to destination sheet first
                next_row = dest_sheet.max_row + 1
                if dest_sheet.cell(row=1, column=1).value is None:  # If sheet is empty
                    next_row = 1
                
                # Cut operation: Copy to destination then delete from source
                dest_sheet.cell(row=next_row, column=1, value=link)
                source_sheet.delete_rows(link_row)
                
                # Save immediately after cut operation
                wb.save(self.excel_path)
                logging.info(f"Cut link {link} from links to {to_sheet}")
            else:
                logging.warning(f"Link {link} not found in links sheet")
                
        except Exception as e:
            logging.error(f"Error cutting link: {str(e)}")
            # Move to inspect sheet if there's an error
            try:
                inspect_sheet = wb['inspect']
                next_row = inspect_sheet.max_row + 1
                if inspect_sheet.cell(row=1, column=1).value is None:
                    next_row = 1
                inspect_sheet.cell(row=next_row, column=1, value=link)
                wb.save(self.excel_path)
                logging.info(f"Moved problematic link {link} to inspect sheet")
            except:
                logging.error(f"Failed to move link {link} to inspect sheet")
        finally:
            wb.close()

    def get_links_from_sheet(self, sheet_name='links'):
        """Get all links from specified sheet"""
        try:
            logging.info(f"Attempting to read sheet: {sheet_name} from {self.excel_path}")
            
            # First check if sheet exists
            wb = load_workbook(self.excel_path)
            if sheet_name not in wb.sheetnames:
                logging.error(f"Sheet '{sheet_name}' not found in Excel file")
                wb.close()
                return []
            wb.close()
            
            # Read Excel file with explicit header=None to not skip first row
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name, header=None)
            logging.info(f"Raw data from sheet: {df.shape[0]} rows")
            
            # Clean the dataframe
            df = df.dropna()
            if df.empty:
                logging.warning(f"No links found in {sheet_name} sheet after cleaning")
                return []
            
            # Get all values from first column, including the first row
            links = [str(link).strip() for link in df[0].tolist() if str(link).strip()]
            
            if links:
                logging.info(f"Found {len(links)} valid links in {sheet_name} sheet")
                logging.debug(f"First few links: {links[:3]}")
            else:
                logging.warning("No valid links found after processing")
            return links
            
        except Exception as e:
            logging.error(f"Error reading links from Excel: {str(e)}")
            logging.exception("Full error details:")
            return []

    def is_sheet_empty(self, sheet_name='links'):
        """Check if a sheet is empty"""
        try:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)
            return df.empty
        except Exception as e:
            logging.error(f"Error checking sheet {sheet_name}: {str(e)}")
            return True 