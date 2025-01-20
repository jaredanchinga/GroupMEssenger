from controller import Controller
from browser_controller import BrowserController
from ixbrowser_local_api import IXBrowserClient
import logging
import time
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from excel_handler import ExcelHandler
import tkinter.messagebox as messagebox

class ExcelController(Controller):
    def __init__(self):
        super().__init__()
        self.excel_handler = None
        # Set the process_links reference
        self.process_links = self._process_links

    def get_message(self):
        """Get message from GUI interface"""
        return super().get_message()

    def move_link_to_sheet(self, link, from_sheet='links', to_sheet='open'):
        """Move a link from one sheet to another in the existing Excel file"""
        wb = load_workbook(self.excel_path)
        
        try:
            # Get source and destination sheets
            source_sheet = wb[from_sheet]
            dest_sheet = wb[to_sheet]
            
            # Find and remove link from source sheet
            link_row = None
            for row in source_sheet.iter_rows(min_row=1, max_col=1):
                if row[0].value == link:
                    link_row = row[0].row
                    break
                
            if link_row:
                # Copy link to destination sheet first
                next_row = dest_sheet.max_row + 1
                dest_sheet.cell(row=next_row, column=1, value=link)
                
                # Then delete from source sheet
                source_sheet.delete_rows(link_row)
                
                # Save workbook
                wb.save(self.excel_path)
                logging.info(f"Moved link from {from_sheet} to {to_sheet}")
            else:
                logging.warning(f"Link not found in {from_sheet} sheet")
                
        except Exception as e:
            logging.error(f"Error moving link between sheets: {str(e)}")
        finally:
            wb.close()

    def get_links_from_sheet(self):
        """Get all links from the links sheet"""
        try:
            df = pd.read_excel(self.excel_path, sheet_name='links')
            if df.empty:
                return []
            return df.iloc[:, 0].tolist()
        except Exception as e:
            logging.error(f"Error reading links from Excel: {str(e)}")
            return []

    def start(self):
        """Override start to initialize Excel handler"""
        if not self.excel_path:
            messagebox.showerror("Error", "Please select an Excel file first")
            return
            
        try:
            self.excel_handler = ExcelHandler(self.excel_path)
            super().start()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to initialize Excel handler: {str(e)}")

    def _process_links(self, message, links_per_profile=5):
        """Actual link processing implementation"""
        if not self.excel_handler:
            raise Exception("Excel handler not initialized")
            
        suspended_profiles = set()
        
        try:
            # Initialize IX Browser client
            client = IXBrowserClient(target='127.0.0.1', port=53200)
            client.show_request_log = False
            
            # Get and sort profiles
            profiles = client.get_profile_list(page=1, limit=100)
            if profiles is None:
                logging.error(f'Error getting profile list: {client.message}')
                return
            
            profile_ids = [profile['profile_id'] for profile in profiles]
            profile_ids.sort()
            
            while True:
                # Check for cancel
                if self.is_cancelled:
                    logging.info("Operation cancelled by user")
                    return
                
                # Check for pause
                while self.is_paused:
                    time.sleep(1)  # Wait while paused
                    if self.is_cancelled:  # Check cancel while paused
                        logging.info("Operation cancelled by user")
                        return
                
                # Get fresh list of remaining links using ExcelHandler
                remaining_links = self.excel_handler.get_links_from_sheet('links')
                if not remaining_links:
                    logging.info("No more links to process")
                    break
                
                total_remaining = len(remaining_links)
                logging.info(f"Current remaining links: {total_remaining}")
                
                # Process each profile in order
                for i, profile_id in enumerate(profile_ids):
                    if self.is_cancelled:
                        return
                    
                    while self.is_paused:
                        time.sleep(1)
                        if self.is_cancelled:
                            return
                    
                    if profile_id in suspended_profiles:
                        logging.info(f"Skipping suspended profile {profile_id}")
                        continue
                    
                    # Get fresh count of remaining links before each profile
                    remaining_links = self.excel_handler.get_links_from_sheet('links')
                    total_remaining = len(remaining_links)
                    if total_remaining == 0:
                        logging.info("No more links to process")
                        return
                    
                    # Calculate batch size - either 5 or whatever is left
                    batch_size = min(links_per_profile, total_remaining)
                    if batch_size == 0:
                        logging.info("No more links to process")
                        return
                    
                    logging.info(f"[{i+1}/{len(profile_ids)}] Profile {profile_id} processing {batch_size} out of {total_remaining} remaining links")
                    
                    # Take only the links we'll process in this batch
                    current_batch = remaining_links[:batch_size]
                    
                    # Process the batch with current profile
                    result = client.open_profile(profile_id, cookies_backup=True, load_extensions=True)
                    if result is None:
                        logging.error(f'Error opening profile {profile_id}')
                        continue
                    
                    browser = BrowserController()
                    links_processed = 0
                    
                    try:
                        browser.initialize_browser(profile_id)
                        
                        # Process each link in current batch
                        for index, link in enumerate(current_batch, 1):
                            logging.info(f"Processing link {index}/{batch_size} with profile {profile_id}")
                            
                            try:
                                result = browser.send_message_to_group(link, message)
                                
                                # Move link using ExcelHandler
                                if result['status'] == 'permanent_skip':
                                    if result['reason'] == 'Profile not logged in':
                                        suspended_profiles.add(profile_id)
                                        break
                                    elif result['reason'] == 'Group requires admin approval':
                                        self.excel_handler.cut_link_to_sheet(link, 'closed')
                                    elif result['reason'] == 'Group has security question':
                                        self.excel_handler.cut_link_to_sheet(link, 'question')
                                elif result['status'] == 'profile_limit':
                                    self.logger.warning(f"Profile {profile_id} has reached its limit, switching to next profile")
                                    break  # Break the link processing loop to switch profiles
                                elif result['status'] == 'success':
                                    self.excel_handler.cut_link_to_sheet(link, 'open')
                                else:
                                    # Handle any other errors by moving to inspect sheet
                                    self.excel_handler.cut_link_to_sheet(link, 'inspect')
                                
                                links_processed += 1
                                
                            except Exception as e:
                                logging.error(f"Error on link {index}/{batch_size}: {str(e)}")
                                
                    finally:
                        if browser:
                            logging.info(f"Closing profile {profile_id} after processing {links_processed}/{batch_size} links")
                            browser.close_browser()
                            close_result = client.close_profile(profile_id)
                            if close_result is None:
                                logging.error(f'Error closing profile {profile_id}')
                            browser = None
                    
                    # Check remaining links after batch using ExcelHandler
                    if self.excel_handler.is_sheet_empty('links'):
                        logging.info("All links processed successfully")
                        return
                    
                    if len(suspended_profiles) == len(profile_ids):
                        logging.error("All profiles are suspended!")
                        return
                        
            logging.info("Processing complete. All links have been moved to appropriate sheets.")
            
        except Exception as e:
            logging.error(f"General error: {str(e)}")

    def run(self):
        """Main execution method"""
        try:
            # Start the GUI
            self.root.mainloop()
        except Exception as e:
            logging.error(f"Error in run method: {str(e)}")
            raise

if __name__ == "__main__":
    app = ExcelController()
    app.run() 